import openai
# from openai import OpenAI
from tqdm import tqdm
import time
import pandas as pd
import os
import json
import re
from openpyxl import load_workbook

def load_json(file_path):
    with open(file_path, 'r') as f:
        data = json.load(f)
    return data

def save_json(data, save_path):
    with open(save_path, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False)

# 旧的调用方式
def completion_old(model, prompt, max_try=1):
    count = 0
    all_responses = []
    openai.api_base = ''
    openai.api_key = ''
    for i in range(max_try):
        try:
            _response = openai.ChatCompletion.create(
                    model=model,
                    messages=[{"role": "system", "content": prompt}],
                    temperature=2,
                    max_tokens=10,
                    top_p=1,
                    frequency_penalty=0,
                    presence_penalty=0,
                    stop=None,
                    # logprobs=40,
                    n=20
                )
            # time.sleep(0.5)
            # print('resp')
            # print(_response)
            all_responses = [_response['choices'][i]['message']['content'] for i in
                                 range(len(_response['choices']))]
            # message = response.choices[0].message.content
            break
        except Exception as e:
            print(e)
            continue
    return all_responses

def completion_client(model, prompt, max_try=1):
    from openai import OpenAI
    client = OpenAI(
        base_url="",
        api_key=''
    )
    count = 0
    message = ''
    for i in range(max_try):
        try:
            response = client.chat.completions.create(
                model=model,
                messages=[
                    {"role": "system", "content": "You are a professional evaluator."},
                    {"role": "user", "content": prompt},
                ],
                temperature=2,
                # max_tokens=5,
                top_p=1,
                frequency_penalty=0,
                presence_penalty=0,
                stop=None,
                # logprobs=40,
                n=20
                # temperature=0.5,
                # max_tokens=1024,
                # n=1,
                # stop=None
            )

            # message = response.choices[0].message.content
            # print(response)
            all_responses = [response.choices[i].message.content for i in
                                 range(len(response.choices))]
            break
        except Exception as e:
            print(e)
            time.sleep(0.5)
            continue
    return all_responses

def parse_output(output):
    matched = re.search("^ ?([\d\.]+)", output)
    if (matched):
        try:
            score = float(matched.group(1))
        except:
            score = -1
    else:
        score = -1
    return score

def Geval_one(p, q):
    prompt = 'You will be given one context and one question ask about information in the context. Your task is to rate the question based on the given context on one metric.\n\n'
    prompt += 'Evaluation Criteria:\n\nAnswerability(1-5) - this dimension is about the answerability of a question based on a context. The question should be able to be answered based on the given context only. The question should be grammatically and semantically correct and cannot ask for information beyond the scope of the context, and the information asked must be consistent with the context.\n\n'
    prompt += 'Evaluation Steps:\n\n1. Read and Understand the Context: Begin by reading the provided context thoroughly to gain a comprehensive understanding of the information given. \n2. Analyze the Question: In the next step, analyze the proposed question. Check its grammar and semantics, ensuring the question is formulated correctly and understandably.\n3. Cross-reference the Context and the Question: Look back and forth between the context and the question to establish if the question is answerable using the information provided in the context only.\n4. Evaluate for Answerability: Based on the correlation between the context and the question, rate the question\'s answerability. This should be on a scale of 1 to 5, where 1 denotes that the question is not answerable based on the context, and 5 indicates that the question is entirely answerable using the provided context.\n\n'
    prompt += 'Example:\n\n'
    # 输入
    context = 'Context:\n\n{}\n\n'.format(p)
    question = 'Question:\n\n{}\n\n'.format(q)
    prompt += context + question
    prompt += 'Evaluation Form (scores ONLY):\n\n- Answerability:'
    # print(prompt)
    # model = 'gpt-4-1106-preview'
    model = 'gpt-3.5-turbo'
    message = completion_old(model, prompt)
    time.sleep(2)
    return message

def Geval(data, json_path):
    new_data = []
    print(len(data))
    for one in tqdm(data):
        new_one = {k:v for k,v in one.items()}
        p = one['passage']
        q = one['prediction']
        responses = Geval_one(p, q)
        all_scores = [parse_output(x) for x in responses]
        all_scores = [x for x in all_scores if x!=-1]
        if len(all_scores) > 0:
            score = sum(all_scores) / len(all_scores)
        else:
            score = -1
        new_one['response'] = responses
        new_one['all_scores'] = all_scores
        new_one['score'] = score
        new_data.append(new_one)
        if len(new_data) % 5 == 0:
            save_json(new_data, json_path)
        if len(new_data) == 1:
            print(new_data[0])
    return new_data
    
def handle_response(data_path):
    wb = load_workbook(data_path)
    ws = wb.active

    every_data_count = 0
    direct_convert_num_count = 0
    no_num_count = 0
    having_a_judging_word_count = 0
    at_the_begining_exists_a_num_count = 0
    else_count = 0

    def my_float(num):
        num = float(num)
        if num>5:
            num = 5
        elif num<0:
            num = 0
        return num
        
    for index,cell in zip(range(1,ws.max_row+1),ws['C']):
        if index==1:
            continue
        datas = cell.value[1:-1].replace("'","").split(",")

        all_scores = []
        for data in datas:
            every_data_count += 1
            data = data.strip()
            try:
                num = my_float(data)
                all_scores.append(num)
                direct_convert_num_count += 1
                #print(num,'---',data)
            except Exception:
                data = data.lower()
                data = re.escape(data)
                
                nums = re.search(r'[0-9]',data)
                if not nums:
                    no_num_count += 1
                    # print('---',data)
                    continue
                
                nums = re.search(r'[0-9]\.?\d*',data)
                if re.search(r'answerability|answerable|answerabilty|rating|answerablility|answerablitilty|answerabilit|score',data):
                    all_scores.append(my_float(nums[0]))
                    having_a_judging_word_count += 1
                    #print(my_float(nums[0]),'---',data)
                    continue
                elif re.search(r'_[0-9]\.?\d_|\([0-9]\.?\d\)|\[[0-9]\.?\d\]',data):
                    all_scores.append(my_float(nums[0]))
                    having_a_judging_word_count += 1
                    #print(my_float(nums[0]),'---',data)
                    continue
                
                nums = re.match(r'[0-9]\.?\d*',data)
                if nums:
                    num = my_float(nums.group())
                    all_scores.append(num)
                    at_the_begining_exists_a_num_count += 1
                    #print(num,'---',data)
                    continue
                nums = re.match(r'"[0-9]\.?\d*',data)
                if nums:
                    num = my_float(nums.group()[1:])
                    all_scores.append(num)
                    at_the_begining_exists_a_num_count += 1
                    #print(num,'---',data)
                    continue
                
                else_count += 1
                print('---',data)
                
        ws['D'+str(index)] = str(all_scores)
        try:
            ws['E'+str(index)] = sum(all_scores)/len(all_scores)
        except Exception:
            ws['E'+str(index)] = -1
            

    print('------[result]------\n',
    every_data_count,'datas in total\n',
    direct_convert_num_count,'datas are directly convert to num\n',
    no_num_count,'datas have no digit and are skipped\n',
    having_a_judging_word_count, 'datas are identified with judging word\n',
    at_the_begining_exists_a_num_count, 'datas have score at the begining\n',
    else_count,'cannot convert\n',
    '--------------------')

    wb.save(data_path)

if __name__ == "__main__":
   
    data = [{
        'passage': '"Bryan Charles Kocis (May 28, 1962 – January 24, 2007), also known as Bryan Phillips, was a director of gay pornographic films and founder of Cobra Video, a gay porn film studio.  Kocis was murdered at his Dallas Township, Pennsylvania home on January 24, 2007; arson was used in an attempt to disguise the circumstances of his death.  Two escorts, Harlow Cuadra and Joseph Kerekes, were charged and convicted for Kocis\' murder and subsequently given a sentence of life imprisonment without any possibility of parole.\nSchoolboy Crush (2004) is a controversial gay pornographic film directed by Bryan Kocis (under the industry name ""Bryan Phillips""), released on Cobra Video, and cast with Brent Everett and Sean Paul Lockhart under the stage name ""Brent Corrigan"".  Corrigan being underage at the time of filming led to legal actions against Phillips and the withdrawal of the film ""Schoolboy Crush"" from the Cobra Video film catalog."',
        'prediction': 'Who was convicted of murdering the director of Schoolboy Crush?'
    }]
    # api
    json_path = './result/json/geval-gpt3turbo-ans.json'
    excel_path = './result/geval-gpt3turbo-anstest.xlsx'
    res = Geval(data, json_path)
    # print(len(res))
    print(res)
    pd.DataFrame(res).to_excel(excel_path, index=False)
    # handle response to score
    handle_response(excel_path)
   
    